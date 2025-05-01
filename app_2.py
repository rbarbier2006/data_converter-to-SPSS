import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

def get_column_letter(n):
    result = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

st.set_page_config(page_title="Participant Data Converter", layout="centered")
st.title("📊 Participant Data Restructuring Tool")

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file:
    start_cell = st.text_input("Enter top-left cell of the table (e.g., B1)", value="B1").strip().upper()
    num_rows = st.number_input("Number of data rows (including time)", min_value=1, step=1)
    num_cols = st.number_input("Number of data columns (including time column)", min_value=1, step=1)

    if st.button("Convert Excel File"):
        wb = load_workbook(uploaded_file, data_only=True)
        first_sheet = wb.sheetnames[0]
        ws_sample = wb[first_sheet]
        start_row, start_col = coordinate_to_tuple(start_cell)

        # Extract headers (skip time column)
        headers = [
            ws_sample.cell(row=start_row, column=start_col + i).value
            for i in range(1, num_cols)
        ]

        # Extract time values (skip header row)
        time_values = [
            ws_sample.cell(row=start_row + i, column=start_col).value
            for i in range(1, num_rows)
        ]

        # Generate column names like "Tcore: 10.0"
        column_names = [f"{header}: {t}" for header in headers for t in time_values]

        # Extract values from all sheets
        data = []
        participant_names = []

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            row_values = []
            for col_offset in range(1, num_cols):
                for row_offset in range(1, num_rows):
                    val = ws.cell(row=start_row + row_offset, column=start_col + col_offset).value
                    row_values.append(val)
            data.append(row_values)
            participant_names.append(sheet)

        # Create final DataFrame
        df = pd.DataFrame(data, columns=column_names)
        df.insert(0, "Participant", participant_names)

        # Save to BytesIO
        output = BytesIO()
        df.to_excel(output, index=False)
        st.success("✅ File processed successfully!")

        st.download_button(
            label="📥 Download Excel File",
            data=output.getvalue(),
            file_name="participant_time_data_flexible.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )